from datetime import date
from django.contrib.auth.models import User
from django.db.models import Q
from rest_framework import viewsets, permissions, status
from rest_framework.decorators import action
from rest_framework.response import Response

from ..models import Domain, Subject, Holiday, Branch, Semester, CollegeLocation, StudentProfile, RegistrationRequest, AttendanceRecord
from ..serializers import (
    DomainSerializer, SubjectSerializer, HolidaySerializer, BranchSerializer, SemesterSerializer, 
    CollegeLocationSerializer, StudentProfileSerializer, RegistrationRequestSerializer
)

class AcademicViewSet(viewsets.ModelViewSet):
    """Base ViewSet for readable-to-all academic metadata"""
    def get_permissions(self):
        if self.action in ['list', 'retrieve']:
            return [permissions.AllowAny()]
        return [permissions.IsAdminUser()]

class DomainViewSet(AcademicViewSet):
    queryset = Domain.objects.all()
    serializer_class = DomainSerializer

class SubjectViewSet(AcademicViewSet):
    queryset = Subject.objects.all()
    serializer_class = SubjectSerializer

    def get_permissions(self):
        # Admins can do anything; authenticated students can list/retrieve;
        # anonymous users cannot access subject list at all
        if self.action in ['list', 'retrieve']:
            return [permissions.IsAuthenticated()]
        return [permissions.IsAdminUser()]

    def get_queryset(self):
        user = self.request.user
        if user.is_staff:
            return Subject.objects.all()
        
        profile = getattr(user, 'studentprofile', None)
        if profile:
            # All-All subjects (no branch/semester restriction) are always visible
            # Restricted subjects are visible only if branch AND semester match (if restricted)
            from django.db.models import Exists, OuterRef, Q
            qs = Subject.objects.filter(
                (Q(branch__isnull=True) & Q(semester__isnull=True)) |           # All-All
                (Q(branch__isnull=True) & Q(semester_id=profile.semester_id)) |  # Any branch, same sem
                (Q(branch_id=profile.branch_id) & Q(semester__isnull=True)) |   # Same branch, any sem
                (Q(branch_id=profile.branch_id) & Q(semester_id=profile.semester_id)) # Exact match
            )
            # For the Mark Attendance dropdown: only return subjects that have
            # at least one active geofence session configured.
            # Admin list view gets all subjects regardless.
            if self.request.query_params.get('active_only') == '1':
                qs = qs.filter(collegelocation__is_active=True).distinct()
            return qs
        return Subject.objects.none()

class HolidayViewSet(viewsets.ModelViewSet):
    queryset = Holiday.objects.all()
    serializer_class = HolidaySerializer
    def get_permissions(self):
        if self.action in ['list', 'retrieve']: return [permissions.AllowAny()]
        return [permissions.IsAdminUser()]

class BranchViewSet(AcademicViewSet):
    queryset = Branch.objects.all()
    serializer_class = BranchSerializer

class SemesterViewSet(AcademicViewSet):
    queryset = Semester.objects.all()
    serializer_class = SemesterSerializer

class CollegeLocationViewSet(viewsets.ModelViewSet):
    queryset = CollegeLocation.objects.all()
    serializer_class = CollegeLocationSerializer
    permission_classes = [permissions.IsAdminUser]

class RegistrationRequestViewSet(viewsets.ModelViewSet):
    queryset = RegistrationRequest.objects.filter(is_processed=False)
    serializer_class = RegistrationRequestSerializer
    permission_classes = [permissions.IsAdminUser]

    @action(detail=True, methods=['post'])
    def approve(self, request, pk=None):
        reg_request = self.get_object()
        user, created = User.objects.get_or_create(
            username=reg_request.registration_no,
            defaults={
                'email': reg_request.email,
                'first_name': reg_request.full_name
            }
        )
        user.password = reg_request.password
        user.save()

        StudentProfile.objects.update_or_create(
            user=user,
            defaults={
                'registration_no': reg_request.registration_no,
                'phone_no': reg_request.phone_no,
                'domain': reg_request.domain,
                'branch': reg_request.branch,
                'semester': reg_request.semester,
                'is_approved': True,
                'approval_date': date.today()
            }
        )
        
        reg_request.delete()
        return Response({"message": "Student approved and created."})

    @action(detail=True, methods=['post'])
    def reject(self, request, pk=None):
        reg_request = self.get_object()
        reg_request.delete()
        return Response({"message": "Request rejected and cleared from system."})

    @action(detail=False, methods=['post'])
    def bulk_action(self, request):
        ids = request.data.get('ids', [])
        action_type = request.data.get('action') # 'approve' or 'reject'
        
        if not ids or action_type not in ['approve', 'reject']:
            return Response({"error": "Invalid payload. Provide list of ids and an action."}, status=status.HTTP_400_BAD_REQUEST)
        
        requests = RegistrationRequest.objects.filter(id__in=ids)
        success_count = 0
        skipped = []
        
        if action_type == 'approve':
            for reg_request in requests:
                try:
                    # 1. Check if phone number is already taken by ANOTHER student
                    existing_profile = StudentProfile.objects.filter(phone_no=reg_request.phone_no).first()
                    if existing_profile and existing_profile.registration_no != reg_request.registration_no:
                        skipped.append(f"{reg_request.registration_no} (Phone {reg_request.phone_no} already exists)")
                        continue

                    # 2. Process Approval
                    user, created = User.objects.get_or_create(
                        username=reg_request.registration_no,
                        defaults={
                            'email': reg_request.email,
                            'first_name': reg_request.full_name
                        }
                    )
                    user.set_password(reg_request.password)
                    user.save()

                    StudentProfile.objects.update_or_create(
                        user=user,
                        defaults={
                            'registration_no': reg_request.registration_no,
                            'phone_no': reg_request.phone_no,
                            'domain': reg_request.domain,
                            'branch': reg_request.branch,
                            'semester': reg_request.semester,
                            'is_approved': True,
                            'approval_date': date.today()
                        }
                    )
                    reg_request.delete()
                    success_count += 1
                except Exception as e:
                    skipped.append(f"{reg_request.registration_no} (Error: {str(e)})")
        else: # reject
            success_count = requests.count()
            requests.delete()
            
        msg = f"Successfully processed {success_count} requests."
        if skipped:
            msg += f" Skipped {len(skipped)} students due to conflicts: {', '.join(skipped[:5])}"
            if len(skipped) > 5: msg += " ..."
            
        return Response({"message": msg, "success_count": success_count, "skipped_count": len(skipped)})

class StudentProfileViewSet(viewsets.ModelViewSet):
    queryset = StudentProfile.objects.all()
    serializer_class = StudentProfileSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_permissions(self):
        if self.action == 'me':
            return [permissions.IsAuthenticated()]
        if self.action in ['list', 'retrieve', 'update', 'partial_update', 'destroy']:
            return [permissions.IsAdminUser()]
        return super().get_permissions()

    def get_queryset(self):
        qs = StudentProfile.objects.all()
        domain_id   = self.request.query_params.get('domain')
        branch_id   = self.request.query_params.get('branch')
        semester_id = self.request.query_params.get('semester')
        if domain_id:   qs = qs.filter(domain_id=domain_id)
        if branch_id:   qs = qs.filter(branch_id=branch_id)
        if semester_id: qs = qs.filter(semester_id=semester_id)
        return qs

    @action(detail=False, methods=['get', 'patch', 'put'])
    def me(self, request):
        profile = getattr(request.user, 'studentprofile', None)
        if not profile:
            return Response({"error": "Profile not found."}, status=status.HTTP_404_NOT_FOUND)
        
        if request.method == 'GET':
            serializer = self.get_serializer(profile)
            return Response(serializer.data)
        
        # Students cannot change registration_no or email
        data = request.data.copy()
        data.pop('registration_no', None)
        data.pop('email', None)
        
        serializer = self.get_serializer(profile, data=data, partial=True)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=False, methods=['get'])
    def analytics(self, request):
        from .analytics import get_student_analytics
        data = get_student_analytics()
        return Response(data)

    @action(detail=True, methods=['get'])
    def subject_stats(self, request, pk=None):
        instance = self.get_object()
        from django.db.models import Q, Exists, OuterRef
        subjects = Subject.objects.filter(
            (Q(branch__isnull=True) & Q(semester__isnull=True)) |           # All-All
            (Q(branch__isnull=True) & Q(semester_id=instance.semester_id)) | # Any branch, same sem
            (Q(branch_id=instance.branch_id) & Q(semester__isnull=True)) |  # Same branch, any sem
            (Q(branch_id=instance.branch_id) & Q(semester_id=instance.semester_id)) # Exact match
        )
        stats = []
        for sub in subjects:
            present = AttendanceRecord.objects.filter(student=instance, subject=sub, status='present', date__gte=instance.approval_date).count()
            absent = AttendanceRecord.objects.filter(student=instance, subject=sub, status='absent', date__gte=instance.approval_date).count()
            total = present + absent
            percentage = (present / total * 100) if total > 0 else 0
            stats.append({
                'subject_id': sub.id,
                'subject_name': sub.name,
                'present': present,
                'absent': absent,
                'total': total,
                'percentage': round(percentage, 2)
            })
        return Response(stats)

    @action(detail=False, methods=['get'], permission_classes=[permissions.IsAdminUser])
    def export_students(self, request):
        """
        Export filtered student list as a styled Excel file.
        Query params: domain (id), branch (id), semester (id)
        """
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        from django.http import HttpResponse
        from django.utils import timezone as tz
        from datetime import datetime
        now_ist = tz.localtime(tz.now())

        domain_id   = request.query_params.get('domain')
        branch_id   = request.query_params.get('branch')
        semester_id = request.query_params.get('semester')

        qs = StudentProfile.objects.select_related(
            'user', 'domain', 'branch', 'semester'
        ).filter(is_approved=True).order_by('branch__name', 'semester__name', 'registration_no')

        if domain_id:   qs = qs.filter(domain_id=domain_id)
        if branch_id:   qs = qs.filter(branch_id=branch_id)
        if semester_id: qs = qs.filter(semester_id=semester_id)



        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Students'

        # Styles
        HEADER_FILL  = PatternFill('solid', fgColor='0F172A')
        ALT_FILL     = PatternFill('solid', fgColor='F8FAFC')
        thin         = Side(style='thin', color='CBD5E1')
        border       = Border(left=thin, right=thin, top=thin, bottom=thin)
        header_font  = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
        title_font   = Font(name='Calibri', bold=True, color='0F172A', size=16)
        cell_font    = Font(name='Calibri', size=10, color='1E293B')
        center       = Alignment(horizontal='center', vertical='center')
        left         = Alignment(horizontal='left',   vertical='center')

        # Build filter label
        filter_parts = []
        if domain_id:
            d = Domain.objects.filter(id=domain_id).first()
            if d: filter_parts.append(f'Domain: {d.name}')
        if branch_id:
            b = Branch.objects.filter(id=branch_id).first()
            if b: filter_parts.append(f'Branch: {b.name}')
        if semester_id:
            s = Semester.objects.filter(id=semester_id).first()
            if s: filter_parts.append(f'Semester: {s.name}')
        filter_label = '  |  '.join(filter_parts) if filter_parts else 'All Students'

        # Title
        ws.merge_cells('A1:I1')
        ws['A1'].value = 'TAP2PRESENT — Student Report'
        ws['A1'].font = title_font
        ws['A1'].alignment = center
        ws.row_dimensions[1].height = 30

        ws.merge_cells('A2:I2')
        ws['A2'].value = f'Filter: {filter_label}   |   Generated: {now_ist.strftime("%d %B %Y, %I:%M %p IST")}'
        ws['A2'].font = Font(name='Calibri', color='64748B', size=9, italic=True)
        ws['A2'].alignment = center
        ws.row_dimensions[2].height = 16

        ws.append([])
        ws.row_dimensions[3].height = 6

        # Headers
        headers = ['#', 'Registration No', 'Full Name', 'Email', 'Phone', 'Domain', 'Branch', 'Semester', 'Approved On']
        ws.append(headers)
        header_row = ws.max_row
        ws.row_dimensions[header_row].height = 22
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=header_row, column=col_idx)
            cell.font = header_font
            cell.fill = HEADER_FILL
            cell.alignment = center
            cell.border = border

        # Data rows
        for i, s in enumerate(qs, start=1):
            row_fill = ALT_FILL if i % 2 == 0 else None
            row_data = [
                i,
                s.registration_no,
                s.user.get_full_name(),
                s.user.email,
                s.phone_no or '—',
                s.domain.name if s.domain else '—',
                s.branch.name if s.branch else '—',
                s.semester.name if s.semester else '—',
                s.approval_date.strftime('%d-%m-%Y') if s.approval_date else '—',
            ]
            ws.append(row_data)
            data_row = ws.max_row
            ws.row_dimensions[data_row].height = 18
            for col_idx, _ in enumerate(row_data, start=1):
                cell = ws.cell(row=data_row, column=col_idx)
                cell.font = cell_font
                cell.alignment = center if col_idx != 3 else left
                cell.border = border
                if row_fill:
                    cell.fill = row_fill

        # Column widths
        col_widths = [5, 18, 28, 32, 16, 20, 18, 18, 16]
        for col_idx, width in enumerate(col_widths, start=1):
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        # Summary row
        ws.append([])
        ws.append(['', f'Total Students: {qs.count()}'])

        import io
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        filename = f'students_{now_ist.strftime("%Y%m%d_%H%M")}.xlsx'
        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response

    def destroy(self, request, *args, **kwargs):
        instance = self.get_object()
        reg_no = instance.registration_no
        user = instance.user
        email = user.email

        # 1. Delete matching Registration Requests
        RegistrationRequest.objects.filter(Q(registration_no=reg_no) | Q(email=email)).delete()
        
        # 2. Delete all Attendance Records for this student
        AttendanceRecord.objects.filter(student=instance).delete()
        
        # 3. Delete the User (Cascades to StudentProfile)
        user.delete()
        
        return Response({"message": "Student and all unique identifiers cleared from database."})
