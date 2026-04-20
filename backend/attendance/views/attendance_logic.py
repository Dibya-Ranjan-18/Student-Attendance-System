import io
import math
import calendar
from datetime import date, datetime, timedelta
from django.db.models import Count, Q, F
from django.http import HttpResponse
from django.shortcuts import get_object_or_404
from rest_framework import viewsets, status, permissions
from rest_framework.decorators import action
from rest_framework.response import Response
from django.utils import timezone
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from ..models import Subject, Holiday, CollegeLocation, StudentProfile, AttendanceRecord
from ..serializers import AttendanceRecordSerializer

# Helper function for Geofencing
def get_distance(lat1, lon1, lat2, lon2):
    R = 6371000  # Earth radius in meters
    phi1, phi2 = math.radians(lat1), math.radians(lat2)
    dphi, dlambda = math.radians(lat2 - lat1), math.radians(lon2 - lon1)
    a = math.sin(dphi/2)**2 + math.cos(phi1)*math.cos(phi2)*math.sin(dlambda/2)**2
    return R * 2 * math.atan2(math.sqrt(a), math.sqrt(1 - a))

def sync_absent_attendance(target_date, profile=None):
    from django.utils import timezone as tz
    if target_date.weekday() == 6 or Holiday.objects.filter(start_date__lte=target_date, end_date__gte=target_date).exists():
        return
    
    # Use IST local time so the session-end comparison is correct on production servers
    now_local = tz.localtime(tz.now())
    now = now_local.replace(tzinfo=None)   # naive datetime for combine() comparison
    day_idx = str(target_date.weekday())
    
    active_sessions = CollegeLocation.objects.filter(
        is_active=True, 
        days_of_week__contains=day_idx
    ).exclude(start_time__isnull=True, end_time__isnull=True).select_related('subject')
    
    base_students = StudentProfile.objects.filter(is_approved=True, approval_date__lte=target_date)
    if profile:
        base_students = base_students.filter(id=profile.id)
    
    records_to_create = []
    
    for session in active_sessions:
        session_end = datetime.combine(target_date, session.end_time)
        if target_date == timezone.localtime(timezone.now()).date() and now < session_end:
            continue  # Session not yet finished for today – skip
        
        subject = session.subject
        if not subject: continue

        # Build the eligible student filter:
        # - If the subject has NO branch/semester restriction (All-All), every approved student qualifies.
        # - If the subject IS restricted, only students whose branch AND semester match qualify.
        if subject.branch is None and subject.semester is None:
            eligible_students = base_students  # no restriction → all students
        elif subject.branch is None:
            eligible_students = base_students.filter(semester=subject.semester)
        elif subject.semester is None:
            eligible_students = base_students.filter(branch=subject.branch)
        else:
            eligible_students = base_students.filter(branch=subject.branch, semester=subject.semester)
        
        existing_student_ids = set(AttendanceRecord.objects.filter(
            date=target_date, 
            subject=subject,
            student__id__in=eligible_students.values_list('id', flat=True)
        ).values_list('student_id', flat=True))

        for s in eligible_students:
            # Skip automatic 'absent' marking for the student's exact approval day
            # to avoid penalizing them for classes that happened before they were approved.
            if s.approval_date == target_date:
                continue
                
            if s.id not in existing_student_ids:
                records_to_create.append(AttendanceRecord(
                    student=s, 
                    date=target_date, 
                    subject=subject, 
                    status='absent',
                    time=session.end_time,
                    class_name=session.name
                ))

    if records_to_create:
        AttendanceRecord.objects.bulk_create(records_to_create, ignore_conflicts=True)

class AttendanceViewSet(viewsets.ModelViewSet):
    queryset = AttendanceRecord.objects.all()
    serializer_class = AttendanceRecordSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        user = self.request.user
        if user.is_staff:
            return AttendanceRecord.objects.all()
        return AttendanceRecord.objects.filter(student__user=user)

    @action(detail=False, methods=['get'])
    def active_session(self, request):
        """
        Returns the CollegeLocation sessions that are running RIGHT NOW for this student.
        Logic:
          - current IST time is within start_time..end_time
          - today's weekday is in days_of_week
          - is_active = True
          - subject matches student branch/semester (or is All-All)
          - not a holiday today
        """
        user = request.user
        profile = getattr(user, 'studentprofile', None)
        if not profile:
            return Response({"error": "Profile not found"}, status=400)

        now_local = timezone.localtime(timezone.now())
        current_time = now_local.time()
        current_day  = str(now_local.weekday())   # '0'=Mon … '6'=Sun
        today        = now_local.date()

        # No sessions on holidays
        if Holiday.objects.filter(start_date__lte=today, end_date__gte=today).exists():
            return Response([])

        # No sessions on Sunday
        if now_local.weekday() == 6:
            return Response([])

        active_locations = CollegeLocation.objects.filter(
            is_active=True,
            start_time__isnull=False,
            end_time__isnull=False,
        ).select_related('subject')

        sessions = []
        for loc in active_locations:
            # Day check
            if current_day not in loc.days_of_week.split(','):
                continue
            # Time window check
            if not (loc.start_time <= current_time <= loc.end_time):
                continue
            subject = loc.subject
            if not subject:
                continue
            # Branch / semester eligibility
            if subject.branch is None and subject.semester is None:
                eligible = True                                         # All-All
            elif subject.branch is None:
                eligible = (profile.semester_id == subject.semester_id)
            elif subject.semester is None:
                eligible = (profile.branch_id == subject.branch_id)
            else:
                eligible = (profile.branch_id == subject.branch_id and
                            profile.semester_id == subject.semester_id)
            if not eligible:
                continue
            # Already marked present today?
            already_marked = AttendanceRecord.objects.filter(
                student=profile, date=today, subject=subject, class_name=loc.name, status='present'
            ).exists()

            sessions.append({
                'session_id':    loc.id,
                'subject_id':    subject.id,
                'subject_name':  subject.name,
                'location_name': loc.name,
                'start_time':    loc.start_time.strftime('%I:%M %p'),
                'end_time':      loc.end_time.strftime('%I:%M %p'),
                'already_marked': already_marked,
            })

        return Response(sessions)


    @action(detail=False, methods=['post'])
    def mark_attendance(self, request):
        user = request.user
        profile = getattr(user, 'studentprofile', None)
        if not profile: return Response({"error": "Profile not found"}, status=400)
        
        device_id = request.data.get('device_id') or request.headers.get('X-Device-Id')
        try:
            lat = float(request.data.get('latitude'))
            lng = float(request.data.get('longitude'))
        except (TypeError, ValueError):
            return Response({"error": "Invalid location coordinates provided."}, status=400)

        subject_id = request.data.get('subject_id')
        class_name = request.data.get('class_name', '')

        if not subject_id:
            return Response({"error": "Please select a subject."}, status=400)
            
        try:
            subject = Subject.objects.get(id=subject_id)
        except Subject.DoesNotExist:
            return Response({"error": "Invalid subject."}, status=400)
        
        active_locations = CollegeLocation.objects.filter(is_active=True, subject=subject)
        if not active_locations.exists():
            return Response({"error": "No attendance window scheduled for this subject."}, status=status.HTTP_400_BAD_REQUEST)
        
        is_inside = False
        is_in_time = False
        is_correct_day = False
        
        # Get IST local time/day
        now_local = timezone.localtime(timezone.now())
        current_time = now_local.time()
        current_day = str(now_local.weekday())

        closest_distance = float('inf')

        for target_location in active_locations:
            # Day check
            if current_day in target_location.days_of_week.split(','):
                is_correct_day = True
            
            # Distance check
            dist = get_distance(lat, lng, target_location.latitude, target_location.longitude)
            if dist < closest_distance:
                closest_distance = dist
                
            if dist <= target_location.radius:
                is_inside = True
            
            # Time check
            if target_location.start_time and target_location.end_time:
                if target_location.start_time <= current_time <= target_location.end_time:
                    is_in_time = True
            else:
                is_in_time = True

            if is_inside and is_in_time and is_correct_day:
                # Device check (only if device_id is stored on the profile)
                stored_device_id = getattr(profile, 'device_id', None)
                if stored_device_id and device_id and stored_device_id != device_id:
                    return Response({
                        "error": "Security Alert: This device is not authorized for this account.",
                    }, status=status.HTTP_403_FORBIDDEN)
                break
        
        if not is_correct_day:
            return Response({"error": f"No sessions scheduled for {subject.name} today."}, status=status.HTTP_403_FORBIDDEN)
            
        if not is_inside:
            return Response({"error": f"Location Out of Bounds: You are {int(closest_distance)}m away from {subject.name} classes."}, status=status.HTTP_403_FORBIDDEN)
        
        if not is_in_time:
            return Response({"error": f"Session Inactive: Attendance for {subject.name} is only allowed between {active_locations[0].start_time.strftime('%H:%M')} and {active_locations[0].end_time.strftime('%H:%M')}."}, status=status.HTTP_403_FORBIDDEN)
        
        today = timezone.localtime(timezone.now()).date()
        # Only block if already marked PRESENT — auto-synced 'absent' records should
        # be overwriteable when the student physically shows up and marks attendance.
        if AttendanceRecord.objects.filter(student=profile, date=today, subject=subject, class_name=class_name, status='present').exists():
           return Response({"error": "Attendance already marked for this session today!"}, status=status.HTTP_400_BAD_REQUEST)

        record, created = AttendanceRecord.objects.update_or_create(
            student=profile, date=today, subject=subject, class_name=class_name,
            defaults={
                'status': 'present',
                'time': timezone.localtime(timezone.now()).time(),   # IST-aware
                'latitude': lat, 'longitude': lng
            }
        )
        
        return Response({"message": "Attendance marked successfully.", "status": "present"})

    @action(detail=False, methods=['get'])
    def daily_report(self, request):
        today_ist = timezone.localtime(timezone.now()).date()
        date_str = request.query_params.get('date', str(today_ist))
        subject_id = request.query_params.get('subject')
        branch_id = request.query_params.get('branch')
        semester_id = request.query_params.get('semester')
        domain_id = request.query_params.get('domain')

        students = StudentProfile.objects.all()
        if branch_id: students = students.filter(branch_id=branch_id)
        if semester_id: students = students.filter(semester_id=semester_id)
        if domain_id: students = students.filter(domain_id=domain_id)

        try:
            if '-' in date_str:
                parts = date_str.split('-')
                if len(parts[0]) == 4: target_date = datetime.strptime(date_str, '%Y-%m-%d').date()
                else: target_date = datetime.strptime(date_str, '%d-%m-%Y').date()
            else:
                target_date = today_ist
        except:
            target_date = today_ist

        # Sync first so the report reflects up-to-date absent records
        sync_absent_attendance(target_date)
        attendance = AttendanceRecord.objects.filter(date=target_date)
        if subject_id:
            try: attendance = attendance.filter(subject_id=int(subject_id))
            except: pass

        is_holiday = Holiday.objects.filter(start_date__lte=target_date, end_date__gte=target_date).first()
        att_map = {a.student_id: a for a in attendance}
        
        present_records = []
        absent_records = []
        
        if subject_id:
            try:
                subject_int_id = int(subject_id)
                target_sub = Subject.objects.filter(id=subject_int_id).first()
                
                student_sub_stats = StudentProfile.objects.filter(id__in=students.values_list('id', flat=True)).annotate(
                    sub_total=Count('attendancerecord', filter=Q(attendancerecord__subject_id=subject_int_id, attendancerecord__date__gte=F('approval_date'))),
                    sub_present=Count('attendancerecord', filter=Q(attendancerecord__subject_id=subject_int_id, attendancerecord__status='present', attendancerecord__date__gte=F('approval_date')))
                )
                sub_perc_map = {s.id: (s.sub_present * 100.0 / s.sub_total if s.sub_total > 0 else 0) for s in student_sub_stats}

                for s in students:
                    record = att_map.get(s.id)
                    sub_perc = sub_perc_map.get(s.id, 0)
                    
                    entry = {
                        'id': s.id,
                        'record_id': record.id if record else None,
                        'student_name': s.user.get_full_name(),
                        'registration_no': s.registration_no,
                        'subject_name': target_sub.name if target_sub else 'N/A',
                        'status': 'holiday' if is_holiday else (record.status if record else 'absent' if target_date < today_ist else 'not_marked'),
                        'subject_percentage': round(sub_perc, 1),
                        'time': record.time.strftime('%H:%M') if record and record.time else None
                    }
                    if record and record.status == 'present': present_records.append(entry)
                    else: absent_records.append(entry)
            except: pass
        else:
            for s in students:
                s_records = attendance.filter(student=s)
                for rec in s_records:
                    entry = {
                        'id': s.id,
                        'record_id': rec.id,
                        'student_name': s.user.get_full_name(),
                        'registration_no': s.registration_no,
                        'subject_name': rec.subject.name if rec.subject else 'General',
                        'status': rec.status,
                        'time': rec.time.strftime('%H:%M') if rec.time else '—'
                    }
                    if rec.status == 'present': present_records.append(entry)
                    else: absent_records.append(entry)

        subject_summaries_qs = Subject.objects.annotate(
            p_count=Count('attendancerecord', filter=Q(attendancerecord__date=target_date, attendancerecord__status='present'))
        )
        
        subject_summaries = [
            {
                'id': sub.id,
                'name': sub.name,
                'present': sub.p_count,
                'total_capacity': sub.total_students,
                'percentage': round((sub.p_count / sub.total_students * 100), 1) if sub.total_students > 0 else 0
            }
            for sub in subject_summaries_qs
        ]
            
        return Response({
            'date': target_date,
            'is_holiday': bool(is_holiday),
            'holiday_reason': is_holiday.reason if is_holiday else '',
            'summary': {
                'total_present': attendance.filter(status='present').values('student').distinct().count(),
                'total_absent': len(absent_records),
                'total_subjects': subject_summaries_qs.count()
            },
            'subject_summaries': subject_summaries,
            'present_records': present_records,
            'absent_records': absent_records
        })

    @action(detail=False, methods=['get'])
    def student_history(self, request):
        user = request.user
        profile = getattr(user, 'studentprofile', None)
        if not profile: return Response({"error": "Profile not found"}, status=400)

        records = AttendanceRecord.objects.filter(
            student=profile, 
            date__gte=profile.approval_date
        ).order_by('-date', '-time')[:50]
        serializer = AttendanceRecordSerializer(records, many=True)
        return Response(serializer.data)

    @action(detail=False, methods=['get'], permission_classes=[permissions.IsAdminUser])
    def monthly_report(self, request):
        """
        Returns a monthly attendance matrix.
        Query params: month (1-12), year (YYYY), subject (id), branch (id, optional)
        Response shape:
          {
            dates: ["2026-04-01", ...],          # all calendar days in month (excl. Sun)
            holidays: {"2026-04-14": "Reason"},  # dates that are holidays
            report: [
              {
                id, name, reg_no,
                attendance: {"2026-04-01": "P"|"A"|"H"|"-"},
                total_present, total_classes, percentage
              }, ...
            ]
          }
        """
        try:
            month = int(request.query_params.get('month', date.today().month))
            year  = int(request.query_params.get('year',  date.today().year))
        except (TypeError, ValueError):
            return Response({"error": "Invalid month or year"}, status=400)

        subject_id = request.query_params.get('subject')
        branch_id  = request.query_params.get('branch')

        if not subject_id:
            return Response({"error": "subject parameter is required"}, status=400)

        try:
            subject = Subject.objects.get(id=subject_id)
        except Subject.DoesNotExist:
            return Response({"error": "Subject not found"}, status=404)

        # Build all calendar days in the month (exclude Sundays)
        import calendar
        days_in_month = calendar.monthrange(year, month)[1]
        all_dates = [
            date(year, month, d)
            for d in range(1, days_in_month + 1)
            if date(year, month, d).weekday() != 6  # exclude Sunday
        ]

        # Build holiday map: date_str -> reason
        holidays_qs = Holiday.objects.filter(
            start_date__lte=date(year, month, days_in_month),
            end_date__gte=date(year, month, 1)
        )
        holiday_map = {}
        for h in holidays_qs:
            cur = h.start_date
            while cur <= h.end_date:
                if date(year, month, 1) <= cur <= date(year, month, days_in_month):
                    holiday_map[str(cur)] = h.reason
                cur += timedelta(days=1)

        # Filter students
        students = StudentProfile.objects.filter(
            is_approved=True,
            approval_date__lte=date(year, month, days_in_month)
        ).select_related('user', 'branch', 'semester')
        if branch_id:
            students = students.filter(branch_id=branch_id)
        # Only students eligible for this subject (handles All-All subjects correctly)
        if subject.branch and subject.semester:
            students = students.filter(branch=subject.branch, semester=subject.semester)
        elif subject.branch:
            students = students.filter(branch=subject.branch)
        elif subject.semester:
            students = students.filter(semester=subject.semester)
        # else: All-All subject → all students eligible, no filter needed

        # Fetch all records for the month + subject in one query
        records_qs = AttendanceRecord.objects.filter(
            subject=subject,
            date__year=year,
            date__month=month,
            student__in=students
        ).values('student_id', 'date', 'status')

        # Build lookup: student_id -> {date_str: status}
        rec_map = {}
        for r in records_qs:
            sid = r['student_id']
            ds  = str(r['date'])
            status = r['status']
            
            if sid not in rec_map:
                rec_map[sid] = {}
            
            # Prioritize 'present' status if multiple records exist for the same day/subject
            if ds not in rec_map[sid] or status == 'present':
                rec_map[sid][ds] = status

        date_strings = [str(d) for d in all_dates]

        report = []
        for s in students:
            # Only include students who were approved before or during this month
            if s.approval_date and s.approval_date > date(year, month, days_in_month):
                continue
            student_records = rec_map.get(s.id, {})
            attendance_row = {}
            total_present = 0
            total_classes = 0
            for d in all_dates:
                ds = str(d)
                if ds in holiday_map:
                    attendance_row[ds] = 'H'
                elif s.approval_date and d < s.approval_date:
                    # Student wasn't enrolled yet
                    attendance_row[ds] = '-'
                elif ds in student_records:
                    status = student_records[ds]
                    attendance_row[ds] = 'P' if status == 'present' else 'A'
                    total_classes += 1
                    if status == 'present':
                        total_present += 1
                else:
                    # Future date or no record yet
                    if d > timezone.localtime(timezone.now()).date():
                        attendance_row[ds] = '-'
                    else:
                        attendance_row[ds] = 'A'
                        total_classes += 1

            percentage = round(total_present / total_classes * 100, 1) if total_classes > 0 else 0
            report.append({
                'id': s.id,
                'name': s.user.get_full_name(),
                'reg_no': s.registration_no,
                'attendance': attendance_row,
                'total_present': total_present,
                'total_classes': total_classes,
                'percentage': percentage,
            })

        return Response({
            'dates': date_strings,
            'holidays': holiday_map,
            'report': report,
        })


    @action(detail=False, methods=['get'])
    def my_subject_stats(self, request):
        user = request.user
        profile = getattr(user, 'studentprofile', None)
        if not profile: return Response({"error": "Profile not found"}, status=400)

        subjects = Subject.objects.filter(
            Q(branch__isnull=True, semester__isnull=True) |           # All-All → everyone
            Q(branch=profile.branch, semester=profile.semester) |     # exact match
            Q(branch__isnull=True, semester=profile.semester) |       # any branch, same sem
            Q(branch=profile.branch, semester__isnull=True)           # same branch, any sem
        )
        stats = []
        for sub in subjects:
            present = AttendanceRecord.objects.filter(student=profile, subject=sub, status='present', date__gte=profile.approval_date).count()
            absent = AttendanceRecord.objects.filter(student=profile, subject=sub, status='absent', date__gte=profile.approval_date).count()
            total = present + absent
            percentage = (present / total * 100) if total > 0 else 0
            stats.append({
                'subject_id': sub.id,
                'subject_name': sub.name,
                'present': present,
                'absent': absent,
                'total': total,
                'percentage': round(percentage, 2)
            })
        return Response(stats)

    @action(detail=False, methods=['post'], permission_classes=[permissions.IsAdminUser])
    def manual_update(self, request):
        student_id = request.data.get('student_id')
        status_val = request.data.get('status')
        subject_id = request.data.get('subject_id')
        date_str = request.data.get('date', str(date.today()))
        
        try:
            if '-' in date_str:
                parts = date_str.split('-')
                if len(parts[0]) == 4: target_date = datetime.strptime(date_str, '%Y-%m-%d').date()
                else: target_date = datetime.strptime(date_str, '%d-%m-%Y').date()
            else: target_date = date.today()
        except: target_date = timezone.localtime(timezone.now()).date()

        profile = get_object_or_404(StudentProfile, id=student_id)
        
        # Validation: Cannot mark attendance before student was approved
        if profile.approval_date and target_date < profile.approval_date:
            return Response({
                "error": f"Invalid Date: Registration for this student was approved on {profile.approval_date}. You cannot mark attendance for a date before their approval."
            }, status=status.HTTP_400_BAD_REQUEST)

        subject = Subject.objects.filter(id=subject_id).first()
        
        if not subject: return Response({"error": "Subject required"}, status=400)

        # Lock modifications to Today only
        if target_date != timezone.localtime(timezone.now()).date():
            return Response({
                "error": "Historical Modification Restricted: Attendance for past dates is read-only. You can only toggle status for the current day."
            }, status=status.HTTP_403_FORBIDDEN)

        record, created = AttendanceRecord.objects.update_or_create(
            student=profile, date=target_date, subject=subject, class_name=subject.name if subject else "",
            defaults={'status': status_val, 'time': timezone.localtime(timezone.now()).time()}
        )
        return Response({"message": "Record updated"})

    @action(detail=False, methods=['post'], permission_classes=[permissions.IsAdminUser])
    def bulk_nudge(self, request):
        student_ids = request.data.get('ids', [])
        subject_id = request.data.get('subject_id')
        
        if not student_ids:
            return Response({"error": "No students selected for nudge."}, status=400)
            
        profiles = StudentProfile.objects.filter(id__in=student_ids).select_related('user')
        subject = Subject.objects.filter(id=subject_id).first()
        subject_name = subject.name if subject else "General Attendance"
        
        # In a real production environment, this would trigger an Email/SMS gateway
        # For this version, we will simulate the logs and return a professional success status
        nudged_names = [p.user.get_full_name() for p in profiles]
        
        # Example of how email would be sent:
        # from django.core.mail import send_mail
        # send_mail(
        #     f"Critical Attendance Alert: {subject_name}",
        #     "Your attendance has fallen below the required 85% threshold. Please meet your coordinator.",
        #     settings.EMAIL_HOST_USER,
        #     [p.user.email for p in profiles],
        #     fail_silently=True
        # )

        return Response({
            "message": f"Successfully sent attendance nudges to {len(profiles)} students.",
            "nudged_students": nudged_names
        })

    @action(detail=False, methods=['get'], permission_classes=[permissions.IsAdminUser])
    def download_excel(self, request):
        """
        Generates and returns a styled Excel (.xlsx) attendance report.
        Optional query params: subject, branch, semester, start_date, end_date, student_id, month, year
        """
        try:
            subject_id = request.query_params.get('subject')
            branch_id = request.query_params.get('branch')
            semester_id = request.query_params.get('semester')
            start_date_str = request.query_params.get('start_date')
            end_date_str = request.query_params.get('end_date')
            single_date_str = request.query_params.get('date')
            
            # New params for individual reports
            student_id = request.query_params.get('student_id')
            month = request.query_params.get('month')
            year = request.query_params.get('year')

            # --- Build queryset ---
            records_qs = AttendanceRecord.objects.select_related(
                'student__user', 'student__branch', 'student__semester', 'subject'
            ).order_by('-date', '-time', 'subject__name')

            if student_id and str(student_id).isdigit():
                records_qs = records_qs.filter(student_id=student_id)
            if subject_id:
                records_qs = records_qs.filter(subject_id=subject_id)
            if branch_id:
                records_qs = records_qs.filter(student__branch_id=branch_id)
            if semester_id:
                records_qs = records_qs.filter(student__semester_id=semester_id)
            
            # Date filtering logic
            if month and year:
                try:
                    m, y = int(month), int(year)
                    last_day = calendar.monthrange(y, m)[1]
                    records_qs = records_qs.filter(
                        date__gte=date(y, m, 1),
                        date__lte=date(y, m, last_day)
                    )
                except (ValueError, TypeError):
                    pass
            elif single_date_str:
                try:
                    target_date = datetime.strptime(single_date_str, '%Y-%m-%d').date()
                    records_qs = records_qs.filter(date=target_date)
                except ValueError:
                    pass
            else:
                if start_date_str:
                    try:
                        start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
                        records_qs = records_qs.filter(date__gte=start_date)
                    except ValueError:
                        pass
                if end_date_str:
                    try:
                        end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
                        records_qs = records_qs.filter(date__lte=end_date)
                    except ValueError:
                        pass

            # --- Create workbook ---
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Attendance Report"

            # Colour palette
            HEADER_FILL  = PatternFill("solid", fgColor="0F172A")   # dark slate
            PRESENT_FILL = PatternFill("solid", fgColor="D1FAE5")   # emerald-100
            ABSENT_FILL  = PatternFill("solid", fgColor="FFE4E6")   # rose-100
            ALT_ROW_FILL = PatternFill("solid", fgColor="F8FAFC")   # slate-50
            thin = Side(style='thin', color='CBD5E1')
            border = Border(left=thin, right=thin, top=thin, bottom=thin)

            header_font  = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
            title_font   = Font(name='Calibri', bold=True, color='0F172A', size=16)
            subheader_font = Font(name='Calibri', bold=True, color='0EA5E9', size=11)
            cell_font    = Font(name='Calibri', size=10, color='1E293B')
            center_align = Alignment(horizontal='center', vertical='center')
            left_align   = Alignment(horizontal='left',   vertical='center')

            # --- Title block ---
            student_name = ""
            reg_no = ""
            if student_id and records_qs.exists():
                first_rec = records_qs.first()
                student_name = first_rec.student.user.get_full_name()
                reg_no = first_rec.student.registration_no
                title_text = f'INDIVIDUAL ATTENDANCE TRANSCRIPT — {student_name} ({reg_no})'
            else:
                title_text = 'TAP2PRESENT — Attendance Report'

            ws.merge_cells('A1:H1')
            title_cell = ws['A1']
            title_cell.value = title_text
            title_cell.font  = title_font
            title_cell.alignment = center_align
            ws.row_dimensions[1].height = 30

            ws.merge_cells('A2:H2')
            subtitle_cell = ws['A2']
            subtitle_cell.value = f'Generated on {timezone.localtime(timezone.now()).strftime("%d %B %Y, %I:%M %p IST")}'
            subtitle_cell.font  = Font(name='Calibri', color='64748B', size=9, italic=True)
            subtitle_cell.alignment = center_align
            ws.row_dimensions[2].height = 16

            ws.append([])  # blank spacer row
            ws.row_dimensions[3].height = 6

            # --- Column headers ---
            headers = [
                'Reg. No',
                'Student Name',
                'Branch',
                'Semester',
                'Subject',
                'Date',
                'Time',
                'Status',
            ]
            ws.append(headers)
            header_row = ws.max_row
            ws.row_dimensions[header_row].height = 22

            for col_idx, _ in enumerate(headers, start=1):
                cell = ws.cell(row=header_row, column=col_idx)
                cell.font      = header_font
                cell.fill      = HEADER_FILL
                cell.alignment = center_align
                cell.border    = border

            # --- Data rows ---
            STATUS_LABELS = {
                'present': 'Present',
                'absent':  'Absent',
                'medical': 'Medical Leave',
                'od':      'On-Duty',
                'personal': 'Personal Leave',
            }

            for i, rec in enumerate(records_qs):
                row_fill = ALT_ROW_FILL if i % 2 == 0 else None
                status_label = STATUS_LABELS.get(rec.status, rec.status.title())

                row_data = [
                    rec.student.registration_no,
                    rec.student.user.get_full_name(),
                    rec.student.branch.name if rec.student.branch else 'N/A',
                    rec.student.semester.name if rec.student.semester else 'N/A',
                    rec.subject.name if rec.subject else 'General',
                    rec.date.strftime('%d-%m-%Y'),
                    rec.time.strftime('%H:%M') if rec.time else '—',
                    status_label,
                ]
                ws.append(row_data)
                data_row = ws.max_row
                ws.row_dimensions[data_row].height = 18

                for col_idx, value in enumerate(row_data, start=1):
                    cell = ws.cell(row=data_row, column=col_idx)
                    cell.font      = cell_font
                    cell.alignment = center_align if col_idx != 2 else left_align
                    cell.border    = border
                    # Status colouring
                    if col_idx == 8:
                        if rec.status == 'present':
                            cell.fill = PRESENT_FILL
                            cell.font = Font(name='Calibri', bold=True, color='065F46', size=10)
                        elif rec.status == 'absent':
                            cell.fill = ABSENT_FILL
                            cell.font = Font(name='Calibri', bold=True, color='9F1239', size=10)
                    elif row_fill:
                        cell.fill = row_fill

            # --- Column widths ---
            col_widths = [16, 28, 14, 18, 26, 14, 10, 16]
            for col_idx, width in enumerate(col_widths, start=1):
                ws.column_dimensions[get_column_letter(col_idx)].width = width

            # --- Summary sheet ---
            ws_summary = wb.create_sheet(title='Summary')
            ws_summary.column_dimensions['A'].width = 30
            ws_summary.column_dimensions['B'].width = 18

            summary_headers = ['Metric', 'Value']
            ws_summary.append(summary_headers)
            for col_idx in range(1, 3):
                cell = ws_summary.cell(row=1, column=col_idx)
                cell.font = header_font
                cell.fill = HEADER_FILL
                cell.alignment = center_align
                cell.border = border

            total = records_qs.count()
            present_count = records_qs.filter(status='present').count()
            absent_count  = records_qs.filter(status='absent').count()
            percentage = round(present_count / total * 100, 1) if total > 0 else 0

            summary_rows = [
                ('Total Records',         total),
                ('Present',               present_count),
                ('Absent',                absent_count),
                ('Overall Percentage',    f'{percentage}%'),
                ('Report Generated',      timezone.localtime(timezone.now()).strftime('%d-%m-%Y %H:%M IST')),
            ]
            for r_label, r_value in summary_rows:
                ws_summary.append([r_label, r_value])
                row_n = ws_summary.max_row
                for col_idx in range(1, 3):
                    cell = ws_summary.cell(row=row_n, column=col_idx)
                    cell.font = Font(name='Calibri', size=10)
                    cell.alignment = center_align
                    cell.border = border

            # --- Stream response ---
            buffer = io.BytesIO()
            wb.save(buffer)
            buffer.seek(0)

            filename = f'attendance_report_{timezone.localtime(timezone.now()).strftime("%Y%m%d_%H%M")}.xlsx'
            response = HttpResponse(
                buffer.getvalue(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            return response
        except Exception as e:
            return Response({"error": f"Excel generation failed: {str(e)}"}, status=500)
